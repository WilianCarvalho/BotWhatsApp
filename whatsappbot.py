
import numpy as np
from selenium import webdriver
import time as tempo
from selenium.webdriver.common.keys import Keys
import pyodbc
import pandas as pd
import urllib
import PySimpleGUI as sg
from datetime import date
from openpyxl import load_workbook

pd.options.mode.chained_assignment = None

data_atual = date.today()

def formatcnpj(dado):
    return "%s.%s.%s/%s-%s" % (dado[0:2], dado[2:5], dado[5:8], dado[8:12], dado[12:])


def formatcpf(cpf):
    return "%s.%s.%s-%s" % (cpf[:3], cpf[3:6], cpf[6:9], cpf[9:])


def formatdata(data):
    return "%s-%s-%s" % (data[0:2], data[2:4], data[4:])
# layout do script
layout = [[sg.Text('Digite a data inicial')],
          [sg.Input(key='datainicial')],
          [sg.Text('Digite a data final')],
          [sg.Input(key='datafinal')],
          [sg.Button('Iniciar o Bot')],

          ]

janela = sg.Window('WhatsBot', layout)
datainicial = ''
datafinal = ''

# Captura os valores digitados nos campos
while True:
    eventos, valores = janela.read()
    if(eventos == sg.WINDOW_CLOSED):
        break
    if(eventos == 'Iniciar o Bot'):
        if(len(valores['datainicial']) >= 8 & len(valores['datafinal']) >= 8):
            datainicial = valores['datainicial']
            datafinal = valores['datafinal']
            break
        else:
            sg.popup('A data esta com o formato errado!!')
# Formata a data da maneira correta
datainicial = formatdata(datainicial)
datafinal = formatdata(datafinal)
# Faz a conexão com o banco de dados
conn = pyodbc.connect(
    'DRIVER={SQL SERVER Native Client 11.0};SERVER=192.168.1.239;DATABASE=sapiens;UID=sapiens;PWD=sapiens')
cursorBdSQL = conn.cursor()
comando = f"""SELECT E301TCR.CODCLI,  
                  E085CLI.NOMCLI,
                  E085CLI.INTNET,
                  E085CLI.CGCCPF,
                  E085CLI.TIPCLI,
                  E085CLI.FONCLI,
                  E085CLI.FONCL2,
                  E070FIL.NOMFIL,
                  E301TCR.CODEMP,
                  E301TCR.CODFIL,
                  E301TCR.VCTPRO,
                  E301TCR.VLRABE,
                  E301TCR.NUMTIT,
            COUNT(E301TCR.NUMTIT) AS QTDTIT
             FROM E301TCR,E085CLI,E070FIL,E002TPT
            WHERE E301TCR.VLRABE > 0
              AND E301TCR.VCTPRO between '{datainicial}' and '{datafinal}'
              AND E002TPT.recsom IN ('D','O')
              AND E085CLI.TIPMER = 'I'
              AND E301TCR.CODTPT IN ('01','02')
              AND E301TCR.CODPOR NOT IN ('DEV','NR','C001','C002')
              AND E301TCR.CODEMP = '01'
              AND E301TCR.CODCLI = E085CLI.CODCLI
              AND E301TCR.CODEMP = E070FIL.CODEMP
              AND E301TCR.CODFIL = E070FIL.CODFIL
              AND E301TCR.CODTPT = E002TPT.CODTPT
              
         GROUP BY E301TCR.CODCLI,
                  E085CLI.NOMCLI,
                  E085CLI.INTNET,
                  E085CLI.CGCCPF,
                  E085CLI.TIPCLI,
                  E085CLI.FONCLI,
                  E085CLI.FONCL2,
                  E070FIL.NOMFIL,
                  E301TCR.CODEMP,
                  E301TCR.CODFIL,
                  E301TCR.VLRABE,
                  E301TCR.NUMTIT,
                  E301TCR.VCTPRO
         ORDER BY E301TCR.CODCLI"""

df = pd.read_sql(comando, conn)
dfauxiliar = pd.read_excel('C:/LogBot/WhatsBotLog.xlsx')



listaUsersSemFon = []
# Inicia a pagina de login do whatsap
navegador = webdriver.Chrome("C:/driver/chromedriver.exe")
navegador.get("https://web.whatsapp.com/")
while len(navegador.find_elements_by_id("side")) < 1:
    tempo.sleep(1)

# trata o telefone que vem do banco
df['FONCL2'] = df['FONCL2'].str.replace(r'[^0-9]', '', regex=True)
df['FONCLI'] = df['FONCLI'].str.replace(r'[^0-9]', '', regex=True)
df['FONCL2'] = '55' + df['FONCL2'].astype(str)
df['FONCLI'] = '55' + df['FONCLI'].astype(str)
# Remove o numero 9 do campo foncli
for i, foncl2 in enumerate(df['FONCL2']):
     if len(df['FONCL2'][i]) >= 13:
         df['FONCL2'][i]=df['FONCL2'][i][:4]+df['FONCL2'][i][5:]
for i, foncli in enumerate(df['FONCLI']):
     if int(df['FONCLI'][i][4]) != 3:
         if len(df['FONCLI'][i]) >= 13:
             df['FONCLI'][i]=df['FONCLI'][i][:4]+df['FONCLI'][i][5:]
df.fillna(0, inplace=True)
dataBuscaFormatadaDia = data_atual.day
dataBuscaFormatadaMes = data_atual.month
dataBuscaFormatadaAno = data_atual.year
dataBuscaFormatada = f'{dataBuscaFormatadaAno}-{dataBuscaFormatadaMes}-{dataBuscaFormatadaDia}'

for i, row in df.iterrows():
     nomcli = row["NOMCLI"]
     telefone = np.int64(row["FONCL2"])
     cnpj = str(row["CGCCPF"])
     dataVencimento = row["VCTPRO"]
     valorDivida = row["VLRABE"]
     numeroTitulo = row["NUMTIT"]
     planilhaNomCli = dfauxiliar.loc[dfauxiliar['nomeCli']==nomcli]
     planilhaData = dfauxiliar.loc[dfauxiliar['dataAtual'] == dataBuscaFormatada]
     planilhaNumtit = dfauxiliar.loc[dfauxiliar['numtit'] == numeroTitulo]
     result = dfauxiliar.loc[(dfauxiliar['nomeCli'] == nomcli) & (dfauxiliar['dataAtual'] == dataBuscaFormatada) & (dfauxiliar['numtit'] == numeroTitulo)]
     if(len(result)==0):
         if int(telefone) < 2:
             telefone = np.int64(row["FONCLI"])
         if int(telefone) > 4:
             if len(str(cnpj)) == 11:
                 cnpjoucpf = formatcpf(str(cnpj))
                 data_em_texto = dataVencimento.strftime('%d/%m/%Y')
             else:
                 cnpjoucpf = formatcnpj(str(cnpj))
                 data_em_texto = dataVencimento.strftime('%d/%m/%Y')
             mensagem = urllib.parse.quote(f"""
                  Prezado cliente {nomcli}  portador do CPF/CNPJ Nº{cnpjoucpf}. Estamos realizando uma conciliação e 
                  identificação dos recebimentos dos nossos clientes, constam titulos vencidos em seu nome. Titulo Nº 
          {numeroTitulo}, com vencimento na data {data_em_texto}, que possui valor de R$ {valorDivida}. Solicito que 
          nos informe se este titulo ja foi pago, e nos envie o comprovante ou a previsão de pagamento do mesmo. Caso o mesmo 
          ja tenha sido liquidado, desconsidere esta mensagem. Ressaltamos que os titulos vencidos são enviados ao SERASA apos 
          o quinto dia de vencimento ou caso se não recebermos um posicionamento da sua parte! 

          Att. Clarice Eletrodomesticos LTDA""")
             url = f"https://web.whatsapp.com/send?phone={telefone}&text={mensagem}"
             navegador.get(url)
             tempo.sleep(20)
             while len(navegador.find_elements_by_id("side")) < 1:
                 tempo.sleep(1)
             # NumeroNãoENCONTRADO
             if len(navegador.find_elements_by_xpath('//*[@id="app"]/div[1]/span[2]/div[1]/span/div[1]/div/div')) != 0:
                 while len(navegador.find_elements_by_id("side")) < 1:
                     tempo.sleep(10)
                 navegador.find_element_by_xpath(
                     '//*[@id="app"]/div[1]/span[2]/div[1]/span/div[1]/div/div/div/div/div[2]/div').send_keys(
                     Keys.ENTER)
                 # COMO O PRIMEIRO NÃO FOI ENCONTRADO É ENVIADO AO 2 NUMERO
                 telefone2 = telefone = np.int64(row["FONCLI"])
                 url = f"https://web.whatsapp.com/send?phone={telefone2}&text={mensagem}"
                 navegador.get(url)
                 # Aguarda o navegadorCarregar
                 while len(navegador.find_elements_by_id("side")) < 1:
                     tempo.sleep(1)
                 tempo.sleep(10)

                 if len(navegador.find_elements_by_xpath(
                         '//*[@id="app"]/div[1]/span[2]/div[1]/span/div[1]/div/div')) == 1:
                     # Se não for ele adiciona aos usuarios não enviados
                     navegador.find_element_by_xpath(
                         '//*[@id="app"]/div[1]/span[2]/div[1]/span/div[1]/div/div/div/div/div[2]/div').send_keys(
                         Keys.ENTER)
                     tempo.sleep(5)
                     listaUsersSemFon.append([nomcli, numeroTitulo])
                 else:
                     # Aguarda o navegador carregar e envia para o segundo numero
                     while len(navegador.find_elements_by_id("side")) < 1:
                         tempo.sleep(1)
                     tempo.sleep(30)
                     wb = load_workbook('C:/LogBot/WhatsBotLog.xlsx')
                     ws = wb.active
                     ws.append({'A': nomcli, 'B': data_atual,'C':numeroTitulo})
                     wb.save('C:/LogBot/WhatsBotLog.xlsx')
                     wb.close()
                     navegador.find_element_by_xpath(
                         '/html/body/div[1]/div[1]/div[1]/div[4]/div[1]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div[2]').send_keys(
                         Keys.ENTER)
                     tempo.sleep(10)
             # NUMERO ENCONTRADO
             else:
                 tempo.sleep(25)
                 wb = load_workbook('C:/LogBot/WhatsBotLog.xlsx')
                 ws = wb.active
                 ws.append({'A': nomcli, 'B': data_atual, 'C': numeroTitulo})
                 wb.save('C:/LogBot/WhatsBotLog.xlsx')
                 wb.close()
                 navegador.find_element_by_xpath(
                     '/html/body/div[1]/div[1]/div[1]/div[4]/div[1]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div[2]').send_keys(
                     Keys.ENTER)
                 tempo.sleep(10)
         else:
             print('Não foi encontrado o numero do cliente ')
             listaUsersSemFon.append([nomcli, numeroTitulo])


# Salva os usuarios que ocorreu um problema para enviar manualmente
sg.popup('Finalizado a Execução!!!!')
tempo.sleep(10)
janela.close()
data_em_texto = data_atual.strftime('%d/%m/%Y')
caminho ='C:/LogBot/logBotWhats.txt'
try:
     arquivo = open(file=caminho, mode='a')
     arquivo.write('Data:'+data_em_texto+'-Esses foram os usuarios que não foi encontrado telefone'+'\n')
     for i in range(len(listaUsersSemFon)):
         for j in range(len(listaUsersSemFon[i])):
             arquivo.write(str(listaUsersSemFon[i][j])+'\n')
     arquivo.close()
except FileNotFoundError:
     arquivo = open(file=caminho, mode='w+')
     arquivo.write('Data:'+data_em_texto+'-Esses foram os usuarios que não foi encontrado telefone'+'\n')
     for i in range(len(listaUsersSemFon)):
         for j in range(len(listaUsersSemFon[i])):
             arquivo.write(str(listaUsersSemFon[i][j])+'\n')
     arquivo.close()